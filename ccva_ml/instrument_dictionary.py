from functools import lru_cache
from pathlib import Path
import json
import re

from openpyxl import load_workbook


PACKAGE_DIR = Path(__file__).resolve().parent
RESOURCE_DIR = PACKAGE_DIR / "resources"
DICTIONARY_DIR = RESOURCE_DIR / "dictionaries"

WORKBOOKS = {
    "2016": RESOURCE_DIR / "va_instr_2016.xlsx",
    "2022": RESOURCE_DIR / "va_instr_2022.xlsx",
}

TARGET_ALIASES = {
    "cod_who_ucod": "pcva_ucod",
    "cod_who_icd": "pcva_ucod_icd",
    "pcva_ucod_icod": "pcva_ucod_icd",
}

EXCLUDED_TYPE_PREFIXES = (
    "begin ",
    "end ",
    "calculate",
    "note",
    "group",
    "repeat",
    "metadata",
)

EXCLUDED_TYPE_EXACT = {
    "acknowledge",
    "audio",
    "barcode",
    "binary",
    "deviceid",
    "end time",
    "end-time",
    "image",
    "instanceid",
    "start time",
    "start-time",
    "time",
}


def normalize_column_name(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s/]+", "_", text)
    text = re.sub(r"[^a-z0-9_]+", "", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def _clean_value(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _first_value(row, *keys):
    for key in keys:
        value = row.get(key)
        if value is not None:
            return _clean_value(value)
    return None


def _extract_choice_list(type_value):
    text = (_clean_value(type_value) or "").lower().replace("_", " ")
    if text.startswith("select one "):
        return _clean_value(text.split(" ", 2)[2])
    if text.startswith("select multiple "):
        return _clean_value(text.split(" ", 2)[2])
    return None


def _is_model_feature(type_value, name):
    normalized_type = (type_value or "").strip().lower().replace("_", " ")
    if not normalized_type:
        return False

    if normalized_type in EXCLUDED_TYPE_EXACT:
        return False

    if any(normalized_type.startswith(prefix) for prefix in EXCLUDED_TYPE_PREFIXES):
        return False

    if normalized_type.startswith("select one") or normalized_type.startswith("select multiple"):
        return True

    allowed_exact = {
        "date",
        "datetime",
        "decimal",
        "integer",
        "phone number",
        "text",
        "time",
        "geopoint",
        "geotrace",
        "geoshape",
    }
    return normalized_type in allowed_exact


def _sheet_rows_as_dicts(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header).strip() if header is not None else None for header in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            if index >= len(row):
                continue
            value = row[index]
            if value is None:
                continue
            record[normalize_column_name(header)] = value
        if record:
            records.append(record)
    return records


def _build_dictionary_from_workbook(workbook_path):
    workbook_path = Path(workbook_path)
    version = "2016" if "2016" in workbook_path.stem else "2022" if "2022" in workbook_path.stem else workbook_path.stem
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    survey_sheet = workbook["survey"]
    choices_sheet = workbook["choices"] if "choices" in workbook.sheetnames else None
    settings_sheet = workbook["settings"] if "settings" in workbook.sheetnames else None

    survey_rows = _sheet_rows_as_dicts(survey_sheet)
    survey = []
    feature_columns = []
    survey_columns = []
    survey_index = {}

    for row in survey_rows:
        name = _clean_value(row.get("name"))
        if not name:
            continue

        normalized_name = normalize_column_name(name)
        type_value = _clean_value(row.get("type"))
        choice_list = _extract_choice_list(type_value)
        include_in_model = _is_model_feature(type_value, name)

        entry = {
            "name": name,
            "normalized_name": normalized_name,
            "type": type_value,
            "label_en": _first_value(row, "labelenglishen", "label_en"),
            "label_sw": _first_value(row, "labelswahilisw", "label_sw"),
            "hint_en": _first_value(row, "hintenglishen", "hint_en"),
            "hint_sw": _first_value(row, "hintswahilisw", "hint_sw"),
            "required": _clean_value(row.get("required")),
            "relevant": _clean_value(row.get("relevant")),
            "agegroup": _clean_value(row.get("agegroup")),
            "choice_filter": _clean_value(row.get("choice_filter")),
            "notes": _clean_value(row.get("notes")),
            "appearance": _clean_value(row.get("appearance")),
            "calculation": _clean_value(row.get("calculation")),
            "default": _clean_value(row.get("default")),
            "constraint": _clean_value(row.get("constraint")),
            "constraint_message_en": _first_value(row, "constraintmessageenglishen", "constraint_message_en"),
            "constraint_message_sw": _first_value(row, "constraintmessageswahilisw", "constraint_message_sw"),
            "read_only": _clean_value(row.get("read_only")),
            "choice_list": choice_list,
            "include_in_model": include_in_model,
        }
        survey.append(entry)
        survey_columns.append(normalized_name)
        survey_index[normalized_name] = entry
        if include_in_model:
            feature_columns.append(normalized_name)

    choice_lists = {}
    if choices_sheet is not None:
        for row in _sheet_rows_as_dicts(choices_sheet):
            list_name = _clean_value(row.get("list_name"))
            option_name = _clean_value(row.get("name"))
            if not list_name or not option_name:
                continue

            normalized_list = normalize_column_name(list_name)
            entry = choice_lists.setdefault(
                normalized_list,
                {
                    "list_name": list_name,
                    "normalized_list_name": normalized_list,
                    "choices": [],
                },
            )
            entry["choices"].append(
                {
                    "name": option_name,
                    "normalized_name": normalize_column_name(option_name),
                    "label_en": _first_value(row, "labelenglishen", "label_en"),
                    "label_sw": _first_value(row, "labelswahilisw", "label_sw"),
                    "region": _clean_value(row.get("region")),
                    "district": _clean_value(row.get("district")),
                }
            )

    settings = {}
    if settings_sheet is not None:
        settings_rows = list(settings_sheet.iter_rows(values_only=True))
        if len(settings_rows) >= 2:
            keys = [normalize_column_name(cell) for cell in settings_rows[0]]
            values = settings_rows[1]
            for index, key in enumerate(keys):
                if not key:
                    continue
                if index < len(values) and values[index] is not None:
                    settings[key] = values[index]

    return {
        "version": version,
        "source_file": str(workbook_path),
        "settings": settings,
        "survey": survey,
        "survey_columns": survey_columns,
        "survey_index": survey_index,
        "feature_columns": feature_columns,
        "choice_lists": choice_lists,
    }


def _dictionary_json_path(version):
    return DICTIONARY_DIR / f"va_instr_{version}.json"


@lru_cache(maxsize=4)
def load_instrument_dictionary(version_or_path):
    path = Path(version_or_path)
    if path.exists():
        if path.suffix.lower() == ".json":
            with open(path, "r", encoding="utf-8") as handle:
                return json.load(handle)
        return _build_dictionary_from_workbook(path)

    version = str(version_or_path).strip().lower()
    if version in {"2016", "2022"}:
        json_path = _dictionary_json_path(version)
        if json_path.exists():
            with open(json_path, "r", encoding="utf-8") as handle:
                return json.load(handle)

        workbook_path = WORKBOOKS[version]
        dictionary = _build_dictionary_from_workbook(workbook_path)
        DICTIONARY_DIR.mkdir(parents=True, exist_ok=True)
        with open(json_path, "w", encoding="utf-8") as handle:
            json.dump(dictionary, handle, indent=2, ensure_ascii=False)
        return dictionary

    raise FileNotFoundError(f"Unsupported instrument dictionary reference: {version_or_path}")


def load_instrument_dictionaries():
    return {version: load_instrument_dictionary(version) for version in WORKBOOKS}


def detect_instrument_version(df, source_path=None, dictionaries=None):
    normalized_columns = {normalize_column_name(column) for column in df.columns}
    source_text = normalize_column_name(Path(source_path).stem) if source_path else ""

    if "2022" in source_text and "2016" not in source_text:
        return {
            "version": "2022",
            "scores": {"2016": 0, "2022": 1},
            "reason": "source_path_hint",
        }

    if "2016" in source_text and "2022" not in source_text:
        return {
            "version": "2016",
            "scores": {"2016": 1, "2022": 0},
            "reason": "source_path_hint",
        }

    dictionaries = dictionaries or load_instrument_dictionaries()
    scores = {}
    for version, dictionary in dictionaries.items():
        survey_overlap = len(normalized_columns.intersection(dictionary.get("survey_columns", [])))
        feature_overlap = len(normalized_columns.intersection(dictionary.get("feature_columns", [])))
        score = survey_overlap + feature_overlap * 2

        if version == "2022":
            if {"pcva_who_cod", "pcva_who_major", "pcva_who_broad"}.intersection(normalized_columns):
                score += 10
            if any(column.startswith("addendum_") for column in normalized_columns):
                score += 5

        if version == "2016" and {"pcva_who_cod", "pcva_who_major", "pcva_who_broad"}.intersection(normalized_columns):
            score -= 2

        scores[version] = score

    best_version = max(scores, key=scores.get)
    return {
        "version": best_version,
        "scores": scores,
        "reason": "column_overlap",
    }


def normalize_target_aliases(df):
    rename_map = {}
    for column in df.columns:
        normalized = normalize_column_name(column)
        if normalized in TARGET_ALIASES:
            rename_map[column] = TARGET_ALIASES[normalized]
    return df.rename(columns=rename_map)


def select_feature_columns(df, dictionary):
    normalized_lookup = {normalize_column_name(column): column for column in df.columns}
    selected = []
    missing = []

    for column in dictionary.get("feature_columns", []):
        original_column = normalized_lookup.get(column)
        if original_column is None:
            missing.append(column)
            continue
        selected.append(original_column)

    return selected, missing
